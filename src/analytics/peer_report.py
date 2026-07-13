import sqlite3
import pandas as pd
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

os.makedirs("output", exist_ok=True)
conn = sqlite3.connect("nifty100.db")
peer = pd.read_sql(
    "SELECT * FROM peer_percentiles",
    conn
)

companies = pd.read_sql(
    "SELECT * FROM companies_clean",
    conn
)

conn.close()

name_col = None

for col in companies.columns:
    if "name" in col.lower():
        name_col = col
        break

if name_col is None:
    name_col = companies.columns[1]

peer = peer.merge(
    companies[["company_id", name_col]],
    on="company_id",
    how="left"
)

peer.rename(
    columns={
        name_col: "company_name"
    },
    inplace=True
)

writer = pd.ExcelWriter(
    "output/peer_comparison.xlsx",
    engine="openpyxl"
)

green = PatternFill(
    fill_type="solid",
    start_color="90EE90"
)

yellow = PatternFill(
    fill_type="solid",
    start_color="FFFF99"
)

red = PatternFill(
    fill_type="solid",
    start_color="FF9999"
)

amber = PatternFill(
    fill_type="solid",
    start_color="FFD966"
)

for group in peer["peer_group_name"].dropna().unique():
    df = peer[
        peer["peer_group_name"] == group
    ].copy()
    df.to_excel(
        writer,
        sheet_name=str(group)[:31],
        index=False
    )

writer.close()

wb = load_workbook(
    "output/peer_comparison.xlsx"
)

for ws in wb.worksheets:
    last_row = ws.max_row
    percentile_col = None

    for cell in ws[1]:
        if cell.value == "percentile_rank":
            percentile_col = cell.column
            break

    if percentile_col:
        for row in range(2, last_row + 1):
            value = ws.cell(
                row=row,
                column=percentile_col
            ).value
            if value is None:
                continue
            if value >= 0.75:
                ws.cell(
                    row=row,
                    column=percentile_col
                ).fill = green
            elif value <= 0.25:
                ws.cell(
                    row=row,
                    column=percentile_col
                ).fill = red
            else:
                ws.cell(
                    row=row,
                    column=percentile_col
                ).fill = yellow

    # Highlight first company
    for cell in ws[2]:
        cell.fill = amber

    # Median row
    ws.append(["Peer Group Median"])

wb.save(
    "output/peer_comparison.xlsx"
)

print("=" * 50)
print("Peer Comparison Report Generated")
print("Saved: output/peer_comparison.xlsx")
print("=" * 50)